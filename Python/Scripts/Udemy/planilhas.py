""" Manipulando planilhas """

import openpyxl

# ! Para abrir a planilha
# pedidos = openpyxl.load_workbook(r'C:\Users\renan\Documents\Programação\VsCode\Python\Scripts\Udemy\Vendas - Dez.xlsx') # noqa

# ! Para acessar as páginas da planilha
# nome = pedidos.sheetnames
# planilha1 = pedidos['Plan1']

# ! Para imprimir os valores da coluna desejada
# for campo in planilha1['c']:
#     print(campo.value)

# ! Para imprimir uma seleção (intervalo) de valores
# for linha in planilha1['a1:c2']:
#     print('', end='\t\t\t')
#     print()
#     for coluna in linha:
#         print(f'{coluna.value}', end='\t\t\t')

# ! Para alterar valores
# for linha in range(5, 16):
#     numero = linha - 1
#     planilha1.cell(linha, 1).value = numero
#     planilha1.cell(linha, 2).value = 1200 + linha

# pedidos.save('nova_planilha.xlsx')
# print('Alteração feita com sucesso.')

# ! Para criar uma nova planilha
planilha = openpyxl.Workbook()
planilha.create_sheet('Planilha1', 0)
planilha.create_sheet('Planilha2', 0)

planilha1 = planilha['Planilha1']
planilha2 = planilha['Planilha2']

# ! Inserindo valores em uma coluna
conta01 = conta02 = 1
for linha in range(1, 11):
    planilha1.cell(linha, 1).value = 500
    planilha1.cell(linha, 2).value = 1200
    planilha2.cell(linha, 1).value = f'Cliente {conta01}'
    conta01 += 1
    planilha2.cell(linha, 2).value = f'Conta {conta02}'
    conta02 += 1
    planilha2.cell(linha, 3).value = 100

# ! Salvando as alterações
planilha.save('abc.xlsx')
print('Salvo com sucesso.')
